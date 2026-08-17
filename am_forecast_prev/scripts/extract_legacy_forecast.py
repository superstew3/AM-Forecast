#!/usr/bin/env python3
"""Extract the legacy forecast series from the Account Manager Sales Dashboard.

Emits a CSV that loads into `legacy_forecast_reference`. Only the months flagged
`promote` become an Original Forecast; the rest are held as a comparison series.

Usage:
    python scripts/extract_legacy_forecast.py <workbook.xlsx> <out.csv>
"""
from __future__ import annotations

import csv
import datetime as dt
import sys
from decimal import Decimal

from openpyxl import load_workbook

SHEET = "Forecast Data"
HEADER_ROW = 4

# Only July 2026 is promoted to Original Forecast. Everything else stays a
# reference series: August to October 2026 overlap the Renewals Pending
# snapshot, which is the authoritative baseline for those months.
PROMOTE_MONTHS = {dt.date(2026, 7, 1)}

# The legacy workbook applies no Highview exclusion. Cameron Stewart's legacy
# forecast totals about $57k across the series against a non-Highview book worth
# roughly $600 a year of actual income, so those lines cannot be shown to be
# Highview-free and are not trusted as a baseline.
UNVERIFIED_MANAGERS = {"Cameron Stewart"}


def australian_fy(d: dt.date) -> int:
    return d.year if d.month >= 7 else d.year - 1


def australian_quarter(d: dt.date) -> int:
    return ((d.month - 7) % 12) // 3 + 1


def extract(path: str):
    wb = load_workbook(path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"sheet '{SHEET}' not found in {path}")
    ws = wb[SHEET]
    rows = []
    for r in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        if not r or r[0] is None:
            continue
        month_raw, source_manager, _reporting, amount = r[0], r[1], r[2], r[3]
        month = (month_raw.date() if isinstance(month_raw, dt.datetime) else month_raw)
        month = month.replace(day=1)
        amount = Decimal(str(amount or 0)).quantize(Decimal("0.01"))
        promote = month in PROMOTE_MONTHS
        clean = source_manager not in UNVERIFIED_MANAGERS
        note = None
        if not clean:
            note = ("Legacy workbook applies no Highview exclusion and this manager's "
                    "legacy values are inconsistent with the non-Highview book. Not used "
                    "as a baseline.")
        rows.append({
            "forecast_month": month.isoformat(),
            "financial_year": australian_fy(month),
            "financial_quarter": australian_quarter(month),
            "source_manager": source_manager,
            "forecast_amount": str(amount),
            "promoted_to_original": promote and clean,
            "is_verified_exclusion_clean": clean,
            "note": note,
        })
    return rows


def main() -> int:
    if len(sys.argv) != 3:
        print(__doc__)
        return 2
    rows = extract(sys.argv[1])
    fields = ["forecast_month", "financial_year", "financial_quarter", "source_manager",
              "forecast_amount", "promoted_to_original", "is_verified_exclusion_clean", "note"]
    with open(sys.argv[2], "w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)

    promoted = [r for r in rows if r["promoted_to_original"]]
    total = sum(Decimal(r["forecast_amount"]) for r in promoted)
    print(f"extracted {len(rows)} legacy forecast rows -> {sys.argv[2]}")
    print(f"promoted to Original Forecast: {len(promoted)} rows, ${total:,.2f}")
    skipped = [r for r in rows
               if r["forecast_month"].startswith("2026-07") and not r["promoted_to_original"]]
    for r in skipped:
        print(f"  July row NOT promoted: {r['source_manager']} ${r['forecast_amount']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
